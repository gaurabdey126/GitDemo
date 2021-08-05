import openpyxl

class HomePageData:

        test_homepage_data = [{"Name":"Gaurab", "Email":"abc@gmail", "Gender":"Male"},{"Name":"Ethan", "Email":"Hunt@gmail", "Gender":"Female"}, {"Name":"James", "Email":"bond@gmail", "Gender":"Female"}]

        @staticmethod
        def getTestData(test_case_name):   ### since static method, self is not required

                Dict = {}
                excel = openpyxl.load_workbook("C:\\Users\\gaura\\Desktop\\Selenium with PYTHON\\Data_for_E2E.xlsx")
                sheet = excel.active

                for i in range(1, sheet.max_row + 1):
                        if sheet.cell(row=i, column=1).value == test_case_name:
                                for j in range(2, sheet.max_column + 1):   ## to get columns
                                        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

                return[Dict]    ### need the data in list, and hence [] around Dict
