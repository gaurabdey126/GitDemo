import openpyxl

excel = openpyxl.load_workbook("C:\\Users\\gaura\\Desktop\\Selenium with PYTHON\\Data_for_E2E.xlsx") ##need to put double slash "\\"
sheet = excel.active   ### to navigate to the active sheet (TC1) in the excel
cell = sheet.cell(row=1, column=2)
# control going from excel -> sheet -> cell

print (cell.value)

# If we want to write something in the excel sheet virtually
sheet.cell(row=2, column=2).value = "James"
print(sheet.cell(row=2, column=2).value)
print(sheet['A5'].value)   ### this will also gives the value like row, column data
print(sheet.max_row)       ### max number of rows in the sheet...excel should be saved
print(sheet.max_column)    ### max number of columns in the sheet...excel should be saved

for i in range (1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "test case 5":   ### suppose we want to print data only for 5th Test Case
        for j in range (1, sheet.max_column+1):              ### to get the column
            print(sheet.cell(row=i, column=j).value)