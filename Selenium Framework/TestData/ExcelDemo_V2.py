import openpyxl

excel = openpyxl.load_workbook("C:\\Users\\gaura\\Desktop\\Selenium with PYTHON\\Data_for_E2E.xlsx")
sheet = excel.active

Dict = {}

for i in range (1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "test case 2":
        for j in range (2, sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value]= sheet.cell(row=i, column=j).value

print(Dict)
