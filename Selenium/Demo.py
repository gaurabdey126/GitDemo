# PIP --->>>> Package Installer for Python.
# You use pip with an install command followed by the name of the package you want to install. pip looks for the package in PyPI, calculates its dependencies, and installs them to ensure requests will work.
# Notice that you use python -m to update pip . The -m switch tells Python to run a module as an executable.

from selenium import webdriver          ### to handle webdriver
import time                             ### to handle "time.sleep()"
driver = webdriver.Firefox(executable_path=)
driver = webdriver.Ie(executable_path=)
driver = webdriver.Chrome(executable_path= "C:\\Users\gaura\Desktop\Selenium with PYTHON\chromedriver.exe")
driver.maximize_window()                #### to maximize the browser
driver.get("https://rahulshettyacademy.com/#/index")
print(driver.title)                     ##### gets the title of the website
print(driver.current_url)               ### to verify if the required website is opened or not
driver.minimize_window()                ### to minize the browser
driver.back()                           #### to navigate back to the main page
driver.refresh()                        ## to referesh the current page
driver.close()                          ### closed the current window where the mouse in pointing
driver.quit()                           ### closes all the window
time.sleep(5)                           ## it will import time
############################################################################################################################################
driver.find_element_by_xpath().send_keys()
driver.find_elements_by_xpath()         #### in group, SEND_KEYS or CLICK will not come up
assert "Sub string" in String           ### validates a value
############################################################################################################################################
dropdown = Select(driver.find_element_by_id("exampleFormControlSelect1")) #Select is a class already present in selenium...but this will work only if the drop down tag ia having SELECT
dropdown.select_by_value_text("Female")      ### it will select the text visible in the dropdown
dropdown.select_by_index(1)             ### it will select the text based on the index number

driver.find_element_by_id("autosuggest").get_attribute('value')   ####.get_attribute is used as the text value goes of on Refreshing
############################################################################################################################################
alert = driver.switch_to.alert ##NO ()..not callable  ### switching the driver to alert and storing this as an alert variable..the driver is in Alert Box
print(alert.text)                       ### getting the text from the alert
assert entertext in alertText           ### validating the entered name is there in the Alert Text or not through Assert
alert.accept()                          ### clicks on positive option (yes, Accept, Ok, etc)
alert.dismiss()                         ### clicks on negative option (no, cancel, etc.)
############################################################################################################################################
# IMPLICIT WAIT
# This is mostly used when we know that the Application that we are Testing is not superfast and there could be delay for few seconds in loading the objects/page
# Basically it set wait for the Driver Object. So that wait is applied GLOBALLY until your test execution is DONE.
# That means this applicable for each and every step of our Test file.
# Beauty of this WAIT is....If the application loads in less that 5 secs, it will not keep waiting and will proceed immediately..MAX wait=5secs
driver.implicitly_wait(7)               ######IMPLICIT wait applied for 5 secs...GLOBALLY
############################################################################################################################################
# EXPLICIT WAIT
# This EXPLICIT wait is used to target only specific element. This will not have effect on any other element/object
# When Appn is superfast and proper load testing is done with SLA n we know only few areas is slow..then use this
#
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

wait = WebDriverWait(driver,5)          ### need to pass 2 arguments...driver itself and wait time in secs and assigning to variable "wait"
wait.until(expected_conditions.presence_of_element_located((By.CLASS_NAME, "promoCode")))  ### .clickable mode and many other are there after expected condition
driver.find_element_by_class_name("promoCode").send_keys("rahulshettyacademy")
driver.find_element_by_css_selector("button.promoBtn").click()

############################################################################################################################################
# list3 = list(filter(None, list2))  ####  to remove extra spaces from list2
# ############################################################################################################################################
# # / vs //:
# A single slash '/' anywhere in XPath signifies to look for the element immediately inside its parent element.
# A double slash '//' signifies to look for any child or any grand-child (descendant) element inside the parent element.
##############################################################################################################################################
# ## Handling MULTIPLE WINDOWS

driver.window_handles
#### this stores all the windows coming up in the page in a LIST and then we have to give indexing to call the proper window
## parent window is always at index[0] and then comes the child window at [1], [2], etc...
parentwindow = driver.window_handles[0]
childwindow1 = driver.window_handles[1]

driver.switch_to.window(childwindow1)  ### switching to child window from parent window with proper argument...now the driver is in childwindow1
print(driver.find_element_by_tag_name("h3").text) ## getting the Text printed
driver.close()

driver.switch_to.window(parentwindow) ### switching back to parent window after the child window is closed....now the driver is in parentwindow
print(driver.find_element_by_tag_name("h3").text)
driver.close()
############################################################################################################################################
# # iframe, frameset, frame tags are mostly used for creating frames on top of HTML
# # Xpath n everything is fine but still why Selenium is not able to identify the object...ANSWER = mostly beacuse FRAME is present and we have to handle frame

#frame id or frame name or index value
driver.switch_to.frame("mce_0_ifr")  ### need to pass argument, which frame and that should we get by using frame id/frame name/frame index value. Used Frame name here
driver.find_element_by_css_selector("#tinymce").clear()
driver.find_element_by_css_selector("#tinymce").send_keys("I am able to AUTOMATE !!!!!!!!!")
driver.switch_to.default_content()  ### it will take the driver to main HTML Page again...this needs to be done or else the driver will keep searching object in the iframe area
#############################################################################################################################################
# # Mouse Hover
### ActionChains is a class that handles mouse hover, rightclick, double click, etc

action = ActionChains(driver)   ### passing the driver as argument...i.e. now action will work on top of driver as an object


menu = driver.find_element_by_id("mousehover")  ### locator step and assigning to a variable that we will pass somewhere
action.move_to_element(menu).perform()   ##### after action., it will just be hover on top of the object identified in menu
## .perform() is always manadtory to make the action works...

childMenu = driver.find_element_by_link_text("Top")  ### locator step and assigning to a variable that we will pass somewhere
##### move to element will make the cursor move to the below element identified by locator
action.move_to_element(childMenu).click().perform() ### .click() as we are clicking the option from the hover option


############################################################################################################################################
# # double click and right click
action = ActionChains(driver)

action.context_click(driver.find_element_by_id("double-click")).perform() ### context_click = right click

action.double_click(driver.find_element_by_id("double-click")).perform() ### double_click = doble click

alert = driver.switch_to.alert

print(alert.text)

alert.accept()

############################################################################################################################################
# # Is Displayed

assert driver.find_element_by_id("displayed-text").is_displayed()

print(driver.find_element_by_id("displayed-text").is_displayed()) ### this is extra just to print the .is_displayed function

driver.find_element_by_id("hide-textbox").click()

assert not driver.find_element_by_id("displayed-text").is_displayed()  ### assert not as the test will not be visible as Hide button is clicked

print(driver.find_element_by_id("displayed-text").is_displayed())

############################################################################################################################################

# # JavaScript Executor: got to console in inspect page and put "document.getElementsByName("name")[0].value"

driver.find_element_by_name("name").send_keys("Hello World!!!")

print(driver.find_element_by_name("name").text)              #### this will not return the extered text

print(driver.find_element_by_name("name").get_attribute("value")) ### this will return the entered text with .get_attribute("value")

print(driver.execute_script('return document.getElementsByName("name")[0].value')) ## this is the JavaScript method when Selenium is not able to get th

########################################################################################################################################
# # Selenium by default cannot handle scroll in a webpage...we need to use JavaScript method for doing that..

print(driver.find_element_by_name("name").get_attribute("value")) ### this will return the entered text with .get_attribute("value")
# ## JavaScript executor for getting the value:
print(driver.execute_script('return document.getElementsByName("name")[0].value')) ## this is the JavaScript method when Selenium is not able to get the value
#
# ##JavaScript executor for clicking an object:
shopButton = driver.find_element_by_css_selector("a[href*='shop']")
driver.execute_script("arguments[0].click();", shopButton)  ### arguments with 0 as there is only one object(shopButton)...if more than one then we have to increase it
#
# ## JavaScript executor for scrolling a page
driver.execute_script("window.scrollTo(0,document.body.scrollHeight);")  #### scrolling the page to the complete page height..top->bottom
driver.execute_script("window.scrollTo(document.body.scrollHeight,0);")  #### scrolling the page to the complete page height..bottom->top






########################################################################################################################################

# Chrome Options:
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--ignore-certificate-errors")
    chrome_options.add_argument("--window-size=1920x1080")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-web-security")
    chrome_options.add_argument("--disable-xss-auditor")
    chrome_options.add_argument("--disable-popup-blocking")
    chrome_options.add_argument("user-agent={userAgent}")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--incognito")

driver = webdriver.Chrome(executable_path="C:\\Users\gaura\Desktop\Selenium with PYTHON\chromedriver.exe", options=chrome_options)
########################################################################################################################################
# Screenshot
driver.get_screenshot_as_file("screen.png")  ## put it in the step we want
#######################################################################################################################################

# PYTEST: py.test/pytest both are fine..but pytest is latest

# any pytest file should start with "test_" or end with "_test"..naming convention for pytest python file -->> it should always start with "test_"

# every code under function(method) and it always should start with "test" keyword...e.g.: def test_firstProgram():

# every CLASS name should start with "Test" for execution...or else it will not get selected

# pytest test runner in pycharm: top right dropdown -> edit configurations -> + -> pytest -> give the file path

# Any code should be wrapped in a method

# through command prompt also we can run all our Pytest Scripts...."cd the path directory for the pytest folder"

# "py.test" --- Then "py.test" command to execute all the scripts present inside that package

# In CMD prompt, we should give "cd folder path" where we want to execute the test case NOT THE PATH OF THE PYTHON FILE

# "py.test -v" --- py.test verbose to see cache directory/Pass/Fail n more details

# "py.test -v -s" --- to see the output of each script in command prompt....by default, it will NOT show the Console Output in cmd prompt

# In PYTEST, every method is treated as ONE TEST CASE.

# In PYTEST, we cannot have methods with same. If we have, then the latest one will overwrite the earlier one.

# "py.test <filename.py> -v -s" --- In PYTEST, if we have to run a single file out of many..."py.test <filename.py> -v -s"....it will execute that particular script.

# Test Case name in PYTEST is method name only..need to write method name very carefully.

# Some common naming covention should be present in the method name...like for credit card test cases.."creditcard" keyword can be put

# "py.test -k <commonkeyword(like creditcard)> -v -s" --- For particular method with a common keyword, we have to put as "py.test -k <commonkeyword(like creditcard)> -v -s"..."k" is used

# -k stands for method execution, -s logs in output, -v stands for more info metadata

# "@pytest.mark.<Name like smoke/sanity>" --- executing few test cases for SMOKE/SANITY like tagging... "@pytest.mark.<Name like smoke/sanity>" on top of the respective methods..need to import "import pytest"

# "py.test -m <marked name> -v -s" --- For marked method execution, "py.test -m <marked name> -v -s" is used..."@pytest.mark.smoke"

# "pytest.mark.skip" --- To skip any particular method, use "pytest.mark.skip"

# "@pytest.mark.xfail" --- For Some methods, we have to execute that scripts but we know, it will fail for data or whatever...but we have to do the exceution...in that case "@pytest.mark.xfail"..it will execute but will not report

#                 # test_demo1.py::test_firstProgram Hello Pytest!!!
#                 # PASSED
#                 # test_demo1.py::test_secondProgram_creditcard Welcome!!!
#                 # XPASS
#                 # test_demo2.py::test_thirdProgram FAILED
#                 # test_demo2.py::test_fourthProgram_creditcard SKIPPED

############################################################################################################################################
# # FIXTURES ---
# Fixtures are functions, which will run before each test function to which it is applied.

# Fixtures are used to feed some data to the tests such as database connections, URLs to test and some sort of input data. ...

# Pytest while the test is getting executed, will see the fixture name as input parameter.

# "@pytest.fixture"" --- this we have to mark...the this execute first and any method calling the method under fixture will be executed after that

# "yield" is a keyword which is used to give the EXIT method...i.e. after after fixture method, calling method and at last what is present in Yield will be executed

# conftest.py --- file to store the fixture for a particular folder...in that folder all the script can utilise it

# fixtures are used as setup and tear down methods for test cases

# "conftest" file to generalise fixture and make it available to all test cases

# Today, we have few test cases and tomorrow we might have "n" no. of test cases...we cannot keep giving fixtures in all method.
# Instead, we can create a class and put all methods under that...n declare "@pytest.mark.usefixtures("Fixture Name")..n then class n undr class all the methods..

# There might be a requirement..where i dont have to open each and every time the browser for each method execution.
# In that scenario, we have to put (scope="class") in conftest file under fixture argument.."@pytest.fixture(scope="class")"

# CONFTEST file ------>>>>>>>
    @pytest.fixture(scope="class")
    def setup():
        driver = webdriver.Chrome(executable_path="C:\\Users\gaura\Desktop\Selenium with PYTHON\chromedriver.exe")
        driver.get("https://rahulshettyacademy.com/angularpractice/")
        print(driver.title)
        yield
        driver.close()
        print("End of the execution")

# # FIXTURE File ---->>>>>>>
    @pytest.mark.usefixtures("setup")
    class Test:
        def test_fixture1(self):
            print("This is the Website for Practice111111")

        def test_fixture2(self):
            print("This is the Website for Practice22222222")

        def test_fixture3(self):
            print("This is the Website for Practice3333333")
############################################################################################################################################

# DATALOAD from FIXTURE:
# We can give the Data in the Conftest file under a fixture and call it in the methods using "@pytest.mark.usefixture("fixture method name")

# Alongwith .self, dataLoad is also passed as ARGUMENT as we are calling DATA in the method...If no data calling was there, then .self is enough

# CONFTEST FILE ----->>>>>>
    import pytest
    @pytest.fixture()
    def data_load():              ### method name should have underscore instead of capital letter
        print("User Profile Data is being created")
        return ["Rahul", "Shetty", "rahulshettyacademy.com"]

# NORMAL METHOD scripts ----->>>>>>
    @pytest.mark.usefixtures("data_load")
    class TestData:
        def test_profile(self, data_load):            ### method name should start with "test_"
            print("First Name: ", data_load[0])
            print("Second Name: ", data_load[1])
            print("Website: ", data_load[2])
#
############################################################################################################################################

# FIXTURE with Multiple Data: Parameterization

# # CONFTEST File ------>>>>>>>
# this fixture is used to call multiple data for a single method
    @pytest.fixture(params=["Chrome", "Mozilla", "IE"]) ### params= is a must and it should hold the multiple data
    def browser(request):                               ### request argument is manadatory
        return request.param                            ### this exact is mandatory

# Normal METHOD script ----->>>>>>
    def test_cross_browser(browser): ## this will run 3 times as we have 3 Data
        print(browser)

# HERE THE SCRIPT WILL RUN 3 TIMES, PICKING UP VALUES FROM PARAM FOR EACH SET OF DATA:
# 1st SET OF DATA = ("Chrome")
# 2nd SET OF DATA = ("Mozilla")
# 3rd SET OF DATA = ("IE")
############################################################################################################################################

# FIXTURE with Multiple Multiple set of data: Parameterization part 2

# this fixture is used to call (multiple+multiple) set of data for a single method

# CONFTEST File ------>>>>>>>
    @pytest.fixture(params=[("Chrome","Ethan", "Hunt"), ("Mozilla","James"), ("IE","Jason")]) ### params= is a must and it should hold the multiple data
    def browser_name(request):                               ### request argument is manadatory
        return request.param

# Normal METHOD script ----->>>>>>
    def test_agent(browser_name):
        print("Browser: ", browser_name[0])
        print("Agent: ", browser_name[1])

# HERE THE SCRIPT WILL RUN 3 TIMES, PICKING UP VALUES FROM PARAM FOR EACH SET OF DATA:
# 1st SET OF DATA = ("Chrome","Ethan", "Hunt")
# 2nd SET OF DATA = ("Mozilla","James")
# 3rd SET OF DATA = ("IE","Jason")
############################################################################################################################################

HTML Report for PYTEST --->>>
# Need to install pytest-html using pip
# pip install pytest-html
#
# "--html=report.html"
# Use command "py.test --html=report.html -v -s"....it will generate in the pytest folder

# then copy the report path from PyCharm in browser to view the report
############################################################################################################################################

LOGGING --->>>

# No need to install anything extra for this Logging File generation.
# <timestamp> : <level> : <testcase name> : <message provide under level>
# Logging Package: Debug, Info, Warning, Error, Critical ---->>>>> this is the hierarchy

import logging  ### this import is for logging
import inspect  ### this import is for inspect

def test_logging_demo():
    loggerName = inspect.stack()[1][3] ### this is used so that, log file have the method name from where it is called

    ### mapping logger, fileHandler, formatter.....same format need to be followed
    # logger = logging.getLogger(__name__)  ## need to import logging..."__name__" it will pull the test case name
    logger = logging.getLogger(loggerName) ### loggerName will pull the class name of the script where it has been used

    fileHandler = logging.FileHandler('logfile_tcname.log') ## logging is mapped to fileHandler which will generate the Log File in the given name

    formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s") ##givin the format for the Log file
    # s is there to capture in string....asctime=time n date, levelname=type of log, name=testcasename, message=msg provided under level name in code

    fileHandler.setFormatter(formatter)  ## passing the formatter in the fileHandler
    logger.addHandler(fileHandler)  ## passing the fileHandler into Logger
    ####### "formatter---->>>>fileHandler---->>>>logger"

    logger.setLevel(logging.INFO)  #### from which hierarchy we need the logs...there can be a requirement where we don't want debug/info..
    ## setting level DEBUG, INFO, WARNING, ERROR, CRITICAL

    #### we can simply put "logging.WARNING" then
    logger.debug("A debug statement is executed")
    logger.info("An info statement is executed")
    logger.warning("WARNING WARNING WARNING")
    logger.error("ERROR ERROR ERROR ERROR")
    logger.critical("SHOWSTOPPER BUG")
############################################################################################################################################

FRAMEWORK --->>>
# Need to install utility package
#
# Create "conftest" and put all the setup stuffs(webdriver invoke, import file, etc) there fixture
#
# Create a class and then a method and under that start writing the selenium code... REMEMBER to call fixture on that class
#
# We can keep the fixture scope to class level if its browser invokation
#
# We should also give the Tear down method..i.e. "yield" which should get executed after the complete code is done
        @pytest.fixture(scope="class")
        def setup(request):    #### request is an instance when fixture is used
            driver = webdriver.Chrome(executable_path="C:\\Users\gaura\Desktop\Selenium with PYTHON\chromedriver.exe")
            driver.maximize_window()
            driver.get("https://rahulshettyacademy.com/angularpractice/")
            request.cls.driver = driver   ### since "return" will not work as "yield" is there... cls.driver is class driver
            ### any class calling this fixture..will have this driver assigned to it...
            ### no need to return anything
            yield
            driver.close()

# In the script, once "cls.driver=driver" is set in fixture..we need to use "self.driver" in any method which is calling that fixture

# We need to create a BaseClass(PARENT CLASS) file under utilities python package...This BaseClass will be called as Inheritance in our reqd method
        @pytest.mark.usefixtures("setup")
        class BaseClass:
            pass

# In reqd test script/method, we have give as below:
        from utilities.BaseClass import BaseClass   ### this import is main
        class TestOne(BaseClass):
            def test_e2e(self):       ### need to pass the fixture method as argument
                self.driver.find_element_by_css_selector("a[href*='shop']").click()
                products = self.driver.find_elements_by_xpath("//div[@class='card h-100']")

# conftest(fixture) ----->>>>> BaseClass(use fixture) ------->>>> reqd class(inheriting BaseClass)

#############################################################################################################################
# invoking browser from command terminal after applying hook:

    def pytest_addoption(parser):
        parser.addoption("--browser_name", action="store", default="chrome") ## commad line settin

    @pytest.fixture(scope="class")
    def setup(request):    #### request is an instance when fixture is used
        browser_name = request.config.getoption("browser_name")   ### this is taking from the command line terminal
            #### BASED ON BROWSER IT WILL SELECT
        if browser_name == "chrome":
            driver = webdriver.Chrome(executable_path="C:\\Users\gaura\Desktop\Selenium with PYTHON\chromedriver.exe")

        elif browser_name == "firefox":
            driver = webdriver.Firefox(executable_path="C:\\Users\gaura\Desktop\Selenium with PYTHON\geckodriver.exe")

        driver.maximize_window()
        driver.get("https://rahulshettyacademy.com/angularpractice/")
        request.cls.driver = driver     ### since "return" will not work as "yield" is there... cls.driver is class driver
                                        ### any class calling this fixture..will have this driver assigned to it...
                                        ### no need to return anything
        yield
        driver.close()

#############################################################################################################################

PAGE OBJECT MODEL --->>>

# Each PAGE will have EACH class and objects will be called from them

# Create a PYTHON PACKAGE and create new new python file in the name of webpages

# conftest:
        PYTHON FILE in the main Framework folder
        It has the driver setup
        Command line terminal browser setup settings
        fixture initialisation with scope='class'
        Need to give the "yield" also

# utilities:
        PYTHON PACKAGE
        It contains the "BaseClass" which calls the fixture from conftest
        BaseClass acts like a PARENT CLASS
        We can add multiple methods which can be REUSED in multiple test cases..like explicit wait,

# pageObjects:
        PYTHON PACKAGE
        It contains the different PYTHON FILE as different pages..
        Different pages are given by us based on our need for Automating an Application..
        In each PAGE, we define the constructor for the driver....
        We keep the XPATH and their respective method..which can be called in our script

# tests:
        PYTHON FILE
        test_<name>.py file...where we write the scripts calling "BaseClass" and driver and objects from the pageObjects files
        Need to assign Object to class from Page Object File..like "checkoutpage = CheckoutPage(self.driver)" and use this as a driver
        And the need to import as below:
                from utilities.BaseClass import BaseClass           ## from <package>.<filename> import <class name>
                from pageObjects.CheckoutPage import CheckoutPage   ## from <package>.<filename> import <class name>

# TestData:
        PYTHON PACKAGE
        Need to create PYTHON FILE for each page (similar to POM) but THIS IS FOR DATA..like HomePageData
        In the script where we are using @pytest.fixture(params = class(Data Page Class). variable (where Data is stored))

# report:
        PYTHON PACKAGE
        All the reports should get generated here.
#############################################################################################################################

# ALLURE REPORT:

# https://mvnrepository.com/artifact/io.qameta.allure/allure-commandline/2.8.1
# Unzip it and put it in some folder(I have put it in Python folder)
# Then add the allure bin path to the PATH OF ENVIRONMENT VARIABLES of system

# #Need to give above methods only
import allure  ## is a must
@allure.description("E2E flow of Buying")
@allure.severity(severity_level="HIGH")
@allure.step("Entering username as {0}")

pytest -v -s --alluredir=reports  ## --alluredir=folder path where u want to stor.....since we are already in the immediate path, hence given the folder name only
allure serve reports  ## to convert the JSON in temporary report
#############################################################################################################################

# PYXL:
import openpyxl

excel = openpyxl.load_workbook("C:\\Users\\gaura\\Desktop\\Selenium with PYTHON\\Data_for_E2E.xlsx")
sheet = excel.active

Dict = {}

for i in range (1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "test case 2":      ###sheet.cell(row=i, column=j).value....  .value is must
        for j in range (1, sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value]= sheet.cell(row=i, column=j).value

print(Dict)


If we want to call a method with CLASS.METHOD NAME...we can only do this for Static Methods...
For all other methods..we have to create OBJECT of that class and then use OBJECT.METHOD NAME
.self paramter is required when the method is NON STATIC..
For @staticmethod, we don't have to declare (self)

